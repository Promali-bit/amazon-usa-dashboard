from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from amazon_reports import StatementData, product_summary


NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "D9EAD3"
RED = "F4CCCC"
AMBER = "FFF2CC"


def _style_sheet(sheet):
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True, size=14)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, 12), 42
        )


def create_excel_report(df: pd.DataFrame, statement: StatementData) -> bytes:
    totals = product_summary(df)
    difference = totals["sales"] - statement.fba_product_sales
    status = "OK" if abs(difference) <= 5 else "REVIEW"

    finance_rows = [
        ["Metric", "Amount"],
        ["FBA product sales", statement.fba_product_sales],
        ["Net income after refunds/credits", statement.income],
        ["Amazon fees and other expenses", statement.expenses],
        ["Net proceeds before transfers", statement.net_before_transfers],
        ["Transfers to bank", statement.transfers],
        ["Remaining after transfers", statement.remaining_after_transfers],
        ["ASIN report sales", totals["sales"]],
        ["Report difference", difference],
        ["Reconciliation status", status],
    ]
    export_columns = {
        "(Child) ASIN": "ASIN",
        "Title": "Product",
        "Sessions - Total": "Sessions",
        "Page Views - Total": "Page Views",
        "Units Ordered": "Units",
        "Total Order Items": "Order Items",
        "Ordered Product Sales": "Sales",
        "Ordered Product Sales - B2B": "B2B Sales",
        "Conversion": "Conversion",
        "Avg. Sale / Unit": "Avg. Sale / Unit",
        "Sales Share": "Sales Share",
        "Revenue / Session": "Revenue / Session",
    }
    detail = df[list(export_columns)].rename(columns=export_columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(finance_rows[1:], columns=finance_rows[0]).to_excel(
            writer, sheet_name="Executive Summary", index=False
        )
        detail.to_excel(writer, sheet_name="ASIN Detail", index=False)
        workbook = writer.book
        summary = workbook["Executive Summary"]
        asin = workbook["ASIN Detail"]

        _style_sheet(summary)
        _style_sheet(asin)
        for row in range(2, summary.max_row + 1):
            summary.cell(row, 2).number_format = "$#,##0.00;[Red]($#,##0.00)"
        summary.cell(summary.max_row, 2).number_format = "@"
        summary.cell(summary.max_row, 2).fill = PatternFill(
            "solid", fgColor=GREEN if status == "OK" else RED
        )
        for row in range(2, asin.max_row + 1):
            for col in (7, 8, 10, 12):
                asin.cell(row, col).number_format = "$#,##0.00"
            for col in (9, 11):
                asin.cell(row, col).number_format = "0.0%"
        for sheet in (summary, asin):
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center")
    return output.getvalue()
